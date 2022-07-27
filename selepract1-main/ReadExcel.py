import openpyxl

path = "C:\Prashanti M\data1.xlsx"
workbook=openpyxl.load_workbook(path)
#sheet = workbook.get_sheet_by_name("Sheet1")
sheet = workbook.active
rows=sheet.max_row
cols=sheet.max_column

print(rows)
print(cols)

for r in range(1,rows+1): #range consider end-1 so we increment
    for c in range(1,cols+1):
        print(sheet.cell(row=r,column=c).value)